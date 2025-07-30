import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.utils import timezone
from .models import Licence, Presence, Session, Ville, Etablissement, User, Profile
from .forms import PresenceForm, SessionForm, LicenceForm, VilleForm, EtablissementForm, GroupForm, UserUpdateForm
from django.utils.timezone import localdate, localtime, now
from datetime import datetime, timedelta
from django.db.models import Count, Max
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import Group, User
from django.contrib import messages
from .forms import UserCreationWithGroupForm
from django.contrib.auth.forms import SetPasswordForm
from django.views.generic.edit import FormView
from django.urls import reverse_lazy, reverse
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin


def accueil(request):
    today = localdate()
    now = localtime()
    sessions_today = Session.objects.filter(date=today)
    sessions = Session.objects.exclude(date=today).order_by('-date')

    return render(request, 'presence/accueil.html', {'sessions_today': sessions_today, 'sessions': sessions, 'today': today, 'now': now,})

def enregistrer_presence(request):
    if request.method == 'POST':
        form = PresenceForm(request.POST)
        if form.is_valid():
            date_aujourdhui = timezone.now().date()
            for licence in Licence.objects.all():
                key = f'licence_{licence.id}'
                present = form.cleaned_data.get(key, False)
                
                # Crée ou met à jour la présence du jour
                Presence.objects.update_or_create(
                    licence=licence,
                    date=date_aujourdhui,
                    defaults={'present': present}
                )
            return redirect('export_presence_du_jour')
    else:
        form = PresenceForm()
    return render(request, 'presence/presence_form.html', {'form': form})

# Gestion des sessions
@login_required
def creer_session(request):
    rapide = request.GET.get('rapide') == '1'  # booléen pour template

    if request.method == 'POST':
        form = SessionForm(request.POST, user=request.user)
        if form.is_valid():
            session = form.save(commit=False)
            session.created_by = request.user

            # 🔧 Établissement = celui du profil de l'utilisateur connecté
            if hasattr(request.user, 'profile'):
                session.etablissement = request.user.profile.etablissement

            session.save()
            form.save_m2m()

            messages.success(request, "Session créée avec succès.")
            return redirect('liste_sessions')
    else:
        initial = {}
        date_from_get = request.GET.get("date")
        heure_debut_from_get = request.GET.get("heure_debut")

        if date_from_get:
            initial["date"] = date_from_get

        if heure_debut_from_get:
            try:
                heure_debut_obj = datetime.strptime(heure_debut_from_get, "%H:%M")
                initial["heure_debut"] = heure_debut_obj.time()
                heure_fin_obj = (heure_debut_obj + timedelta(hours=2)).time()
                initial["heure_fin"] = heure_fin_obj
            except ValueError:
                pass

        form = SessionForm(initial=initial, user=request.user)

    return render(request, 'presence/creer_session.html', {'form': form, 'rapide': rapide})


@login_required
def liste_sessions(request):
    profile = getattr(request.user, 'profile', None)
    etablissement = profile.etablissement if profile else None

    if etablissement:
        sessions_today = Session.objects.filter(date=datetime.today(), etablissement=etablissement)
        sessions_futures = Session.objects.filter(date__gt=datetime.today(), etablissement=etablissement).order_by('date')
        sessions_passees = Session.objects.filter(date__lt=datetime.today(), etablissement=etablissement).order_by('-date')
    else:
        # Si pas d'établissement, on peut afficher aucune session ou toutes, selon ta logique
        sessions_today = Session.objects.none()
        sessions_futures = Session.objects.none()
        sessions_passees = Session.objects.none()

    context = {
        'sessions_today': sessions_today,
        'sessions_futures': sessions_futures,
        'sessions_passees': sessions_passees,
        'today': datetime.today(),
    }
    return render(request, 'presence/liste_sessions.html', context)

def voir_session(request, pk):
    session = get_object_or_404(Session, pk=pk)
    return render(request, 'presence/voir_session.html', {'session': session})

@login_required
def modifier_session(request, pk):
    session = get_object_or_404(Session, pk=pk)

    session_datetime = datetime.combine(session.date, session.heure_debut)
    if timezone.is_naive(session_datetime):
        session_datetime = timezone.make_aware(session_datetime)

    if now() > session_datetime + timedelta(hours=24):
        messages.error(request, "Impossible de modifier une session de plus de 24h.")
        return redirect('liste_sessions')

    if request.method == 'POST':
        form = SessionForm(request.POST, instance=session, user=request.user)
        if form.is_valid():
            session = form.save(commit=False)
            session.checked_by = request.user
            session.save()
            form.save_m2m()
            return redirect('liste_sessions')
    else:
        form = SessionForm(instance=session, user=request.user)

    return render(request, 'presence/modifier_session.html', {'form': form, 'session': session})

@login_required
def check_rapide(request):
    today = localdate()
    now_time = localtime().time()
    now_datetime = datetime.combine(today, now_time)
    heure_fin = (now_datetime + timedelta(hours=2)).time()

    profile = getattr(request.user, 'profile', None)
    etablissement = profile.etablissement if profile else None

    # Création de la session rapide en base
    session = Session.objects.create(
        date=today,
        heure_debut=now_time,
        heure_fin=heure_fin,
        created_by=request.user,
        checked_by=request.user,
        etablissement=etablissement,
    )

    # Redirection avec paramètre rapide
    url = (
        reverse('creer_session') +
        f"?date={today.strftime('%Y-%m-%d')}&heure_debut={now_time.strftime('%H:%M')}&rapide=1"
    )
    return redirect(url)

@login_required
def confirmer_suppression_session(request, pk):
    session = get_object_or_404(Session, pk=pk)

    # Calcul de la date/heure de début de session
    session_datetime = datetime.combine(session.date, session.heure_debut)

    # Rendre timezone-aware si nécessaire
    if timezone.is_naive(session_datetime):
        session_datetime = timezone.make_aware(session_datetime)

    # Bloquer suppression si plus de 24h
    if timezone.now() > session_datetime + timedelta(hours=24):
        messages.error(request, "Impossible de supprimer une session de plus de 24h.")
        return redirect('liste_sessions')

    # Suppression si autorisée
    if request.method == "POST":
        session.delete()
        messages.success(request, "La session a bien été supprimée.")
        return redirect('liste_sessions')

    return render(request, 'presence/confirmer_suppression_session.html', {'session': session})

# Licencier
@login_required
def ajouter_licencie(request):
    if request.method == 'POST':
        form = LicenceForm(request.POST)
        if form.is_valid():
            licence = form.save(commit=False)

            # Associer automatiquement l'établissement depuis le profil utilisateur
            profile = getattr(request.user, 'profile', None)
            if profile and profile.etablissement:
                licence.etablissement = profile.etablissement

            licence.save()
            messages.success(request, "Le licencié a été ajouté avec succès.")
            return redirect('liste_licencies')
    else:
        form = LicenceForm()

    return render(request, 'presence/ajouter_licencie.html', {'form': form})

@login_required
def liste_licencies(request):
    # On part de toutes les licences avec les annotations
    licencies = Licence.objects.annotate(
        nb_presences=Count('sessions'),
        last_session_date=Max('sessions__date')
    )

    etablissements = None
    if request.user.is_superuser:
        etablissements = Etablissement.objects.all()
        etablissement_id = request.GET.get('etablissement')
        if etablissement_id:
            licencies = licencies.filter(etablissement_id=etablissement_id)
    else:
        # Pour les autres utilisateurs, on filtre sur leur établissement lié au profil
        if hasattr(request.user, "profile") and request.user.profile.etablissement:
            licencies = licencies.filter(etablissement=request.user.profile.etablissement)

    return render(request, 'presence/liste_licencies.html', {
        'licencies': licencies,
        'etablissements': etablissements,
    })

@login_required
def detail_licencie(request, licencie_id):
    # Annoter avec la dernière session par rapport au nom correct de la relation
    licencie = Licence.objects.annotate(
        last_session_date=Max('sessions__date')  # <-- correction ici
    ).filter(pk=licencie_id).first()

    # Protection d'accès si non superuser
    if not licencie:
        return render(request, '404.html', status=404)  # En cas de non existence

    if not request.user.is_superuser:
        user_etab = getattr(request.user.profile, 'etablissement', None)
        if user_etab and licencie.etablissement != user_etab:
            return render(request, '403.html')  # Ou: return HttpResponseForbidden()

    return render(request, 'presence/detail_licencie.html', {
        'licencie': licencie
    })

@login_required
def modifier_licencie(request, licencie_id):
    licencie = get_object_or_404(Licence, pk=licencie_id)

    if request.method == 'POST':
        form = LicenceForm(request.POST or None, instance=licencie, user=request.user)
        if form.is_valid():
            form.save()
            return redirect('liste_licencies')
    else:
        form = LicenceForm(instance=licencie)

    return render(request, 'presence/modifier_licencie.html', {'form': form, 'licencie': licencie})

@login_required
def supprimer_licencie(request, licencie_id):
    licencie = get_object_or_404(Licence, pk=licencie_id)

    if request.method == 'POST':
        licencie.delete()
        messages.success(request, f"Licencié {licencie.prenom} {licencie.nom} supprimé.")
        return redirect('liste_licencies')

    return render(request, 'presence/confirmer_suppression_licencie.html', {'licencie': licencie})

# Export data
def export_page(request):
    context = {
        'total_licencies': Licence.objects.count(),
        'total_sessions': Session.objects.count(),
        'total_users': User.objects.filter(is_superuser=False).count(),
        'total_etablissements': Etablissement.objects.count(),
    }
    return render(request, 'presence/export_page.html', context)

@login_required
def export_licencies_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Licenciés"

    current_time = localtime(now())
    export_date = current_time.strftime('%d/%m/%Y')
    export_time = current_time.strftime('%H:%M:%S')
    user = request.user

    # Styles améliorés
    # Couleurs de l'identité visuelle
    primary_color = "2E5BBA"  # Bleu principal
    secondary_color = "4A90A4"  # Bleu-vert
    accent_color = "E8F2FF"  # Bleu très clair
    text_dark = "2C3E50"  # Gris foncé
    text_light = "7F8C8D"  # Gris clair

    # Polices et styles
    title_font = Font(name="Calibri", bold=True, size=16, color=primary_color)
    subtitle_font = Font(name="Calibri", bold=True, size=12, color=text_dark)
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    data_font = Font(name="Calibri", size=10, color=text_dark)
    info_font = Font(name="Calibri", size=9, color=text_light)
    
    # Remplissages
    header_fill = PatternFill("solid", fgColor=primary_color)
    accent_fill = PatternFill("solid", fgColor=accent_color)
    alternate_fill = PatternFill("solid", fgColor="F8FAFC")  # Gris très clair pour lignes alternées
    
    # Bordures
    thick_border = Border(
        left=Side(style='medium', color=primary_color),
        right=Side(style='medium', color=primary_color),
        top=Side(style='medium', color=primary_color),
        bottom=Side(style='medium', color=primary_color)
    )
    thin_border = Border(
        left=Side(style='thin', color='D5DBDB'),
        right=Side(style='thin', color='D5DBDB'),
        top=Side(style='thin', color='D5DBDB'),
        bottom=Side(style='thin', color='D5DBDB')
    )
    
    # Alignements
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    # Configuration des colonnes
    column_config = {
        'B': {'width': 22, 'header': 'Nom'},
        'C': {'width': 22, 'header': 'Prénom'},
        'D': {'width': 18, 'header': 'Grade'},
        'E': {'width': 15, 'header': 'Participations'},
        'F': {'width': 20, 'header': 'Dernière présence'}
    }

    # Fonctions utilitaires
    def merge_and_style(row, col_start, col_end, text, font, fill=None, alignment=center_align):
        if col_start == col_end:
            cell = ws.cell(row=row, column=col_start)
        else:
            ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
            cell = ws.cell(row=row, column=col_start)
        
        cell.value = text
        cell.font = font
        cell.alignment = alignment
        if fill:
            cell.fill = fill
        return cell

    def apply_border_range(start_row, end_row, start_col, end_col, border):
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).border = border

    # Récupérer les informations utilisateur
    profile = getattr(user, 'profile', None)
    etablissement = profile.etablissement.name if profile and profile.etablissement else "Non défini"
    role = user.groups.first().name if user.groups.exists() else "Aucun rôle"

    # En-tête du document (lignes 1-6)
    # Logo/Titre principal
    merge_and_style(1, 2, 6, "RAPPORT DE SUIVI", title_font, accent_fill)
    merge_and_style(2, 2, 6, "Présences des Licenciés", subtitle_font)

    # Informations contextuelles (ligne 4)
    ws.cell(row=4, column=2, value="Exporté par :").font = info_font
    ws.cell(row=4, column=3, value=f"{user.get_full_name() or user.username}").font = data_font
    
    ws.cell(row=4, column=5, value="Date :").font = info_font
    ws.cell(row=4, column=5).alignment = right_align
    ws.cell(row=4, column=6, value=export_date).font = data_font
    ws.cell(row=4, column=6).alignment = right_align

    ws.cell(row=5, column=2, value="Rôle :").font = info_font
    ws.cell(row=5, column=3, value=role).font = data_font
    
    ws.cell(row=5, column=5, value="Heure :").font = info_font
    ws.cell(row=5, column=5).alignment = right_align
    ws.cell(row=5, column=6, value=export_time).font = data_font
    ws.cell(row=5, column=6).alignment = right_align

    ws.cell(row=6, column=2, value="Établissement :").font = info_font
    ws.cell(row=6, column=3, value=etablissement).font = data_font

    # Ligne de séparation (ligne 7)
    for col in range(2, 7):
        cell = ws.cell(row=7, column=col)
        cell.fill = PatternFill("solid", fgColor=primary_color)

    # En-têtes du tableau (ligne 9)
    header_row = 9
    for col_letter, config in column_config.items():
        col_num = ord(col_letter) - ord('A') + 1
        cell = ws.cell(row=header_row, column=col_num, value=config['header'])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thick_border

    # Récupérer et filtrer les licenciés selon les permissions
    licencies = Licence.objects.annotate(
        nb_presences=Count('sessions'),
        last_session_date=Max('sessions__date')
    )

    # Filtrage selon l'établissement de l'utilisateur
    if not request.user.is_superuser:
        if hasattr(request.user, "profile") and request.user.profile.etablissement:
            licencies = licencies.filter(etablissement=request.user.profile.etablissement)

    # Tri alphabétique par nom puis prénom
    licencies = licencies.order_by('nom', 'prenom')

    # Données du tableau (à partir de la ligne 10)
    data_start_row = 10
    for i, licencie in enumerate(licencies):
        row_num = data_start_row + i
        is_alternate = i % 2 == 1
        
        # Données
        data = [
            licencie.nom,
            licencie.prenom,
            licencie.grade,
            licencie.nb_presences,
            licencie.last_session_date.strftime("%d/%m/%Y") if licencie.last_session_date else "Aucune"
        ]
        
        for j, (col_letter, value) in enumerate(zip(column_config.keys(), data)):
            col_num = ord(col_letter) - ord('A') + 1
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = data_font
            cell.border = thin_border
            
            # Alignement selon le type de données
            if col_letter in ['D', 'E']:  # Grade et Participations
                cell.alignment = center_align
            elif col_letter == 'F':  # Date
                cell.alignment = center_align
            else:  # Nom et Prénom
                cell.alignment = left_align
            
            # Couleur de fond alternée
            if is_alternate:
                cell.fill = alternate_fill

    # Ligne de résumé (après les données)
    summary_row = data_start_row + len(licencies) + 1
    
    # Ligne de séparation avant le résumé
    for col in range(2, 7):
        ws.cell(row=summary_row - 1, column=col).fill = PatternFill("solid", fgColor="BDC3C7")
    
    # Résumé
    merge_and_style(summary_row, 2, 3, "TOTAL", Font(bold=True, size=11, color=primary_color), 
                   accent_fill, right_align)
    
    total_cell = ws.cell(row=summary_row, column=4, value=f"{len(licencies)} licencié(s)")
    total_cell.font = Font(bold=True, size=11, color=primary_color)
    total_cell.alignment = center_align
    total_cell.fill = accent_fill
    total_cell.border = thin_border

    total_participations = sum(licencie.nb_presences for licencie in licencies)
    part_cell = ws.cell(row=summary_row, column=5, value=f"{total_participations}")
    part_cell.font = Font(bold=True, size=11, color=primary_color)
    part_cell.alignment = center_align
    part_cell.fill = accent_fill
    part_cell.border = thin_border

    # Pied de page
    footer_row = summary_row + 3
    merge_and_style(footer_row, 2, 6, 
                   f"Document généré automatiquement le {export_date} à {export_time}", 
                   Font(size=8, italic=True, color=text_light))

    # Configuration des largeurs de colonnes
    for col_letter, config in column_config.items():
        ws.column_dimensions[col_letter].width = config['width']

    # Marges et mise en page
    ws.column_dimensions['A'].width = 2  # Marge gauche
    ws.column_dimensions['G'].width = 2  # Marge droite

    # Propriétés de la page
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.7
    ws.page_margins.right = 0.7
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75

    # En-tête et pied de page pour l'impression
    ws.oddHeader.center.text = "Suivi de Présence - Licenciés"
    ws.oddHeader.center.font = "Calibri,Bold"
    ws.oddFooter.center.text = "Page &P sur &N"

    # Génération du nom de fichier
    etablissement_clean = etablissement.replace(" ", "_").replace("/", "-")
    filename = f"Licencies_{etablissement_clean}_{current_time.strftime('%d%m%Y_%H%M')}.xlsx"

    # Réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
def export_donnees_excel(request, mode='all'):
    """
    mode : 'all' | 'users' | 'etablissement'
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Export"

    current_time = localtime(now())
    export_date = current_time.strftime('%d/%m/%Y')
    export_time = current_time.strftime('%H:%M:%S')
    user = request.user
    now_str = current_time.strftime('%Y%m%d_%H%M%S')
    filename = f"Export_{mode}_{now_str}.xlsx"

    # Styles améliorés - cohérents avec les autres exports
    # Couleurs de l'identité visuelle
    primary_color = "2E5BBA"  # Bleu principal
    secondary_color = "4A90A4"  # Bleu-vert
    accent_color = "E8F2FF"  # Bleu très clair
    text_dark = "2C3E50"  # Gris foncé
    text_light = "7F8C8D"  # Gris clair
    header_blue = "4472C4"  # Bleu des en-têtes (pour compatibilité)
    white_text = "FFFFFF"   # Texte blanc
    light_gray = "F2F2F2"   # Lignes alternées
    black_text = "000000"   # Texte noir
    
    # Polices et styles
    title_font = Font(name="Calibri", bold=True, size=16, color=primary_color)
    subtitle_font = Font(name="Calibri", bold=True, size=12, color=text_dark)
    header_font = Font(name="Calibri", bold=True, size=11, color=white_text)
    data_font = Font(name="Calibri", size=10, color=text_dark)
    info_font = Font(name="Calibri", size=9, color=text_light)
    
    # Remplissages
    header_fill = PatternFill("solid", fgColor=primary_color)
    accent_fill = PatternFill("solid", fgColor=accent_color)
    alternate_fill = PatternFill("solid", fgColor="F8FAFC")  # Gris très clair pour lignes alternées
    
    # Bordures
    thick_border = Border(
        left=Side(style='medium', color=primary_color),
        right=Side(style='medium', color=primary_color),
        top=Side(style='medium', color=primary_color),
        bottom=Side(style='medium', color=primary_color)
    )
    thin_border = Border(
        left=Side(style='thin', color='D5DBDB'),
        right=Side(style='thin', color='D5DBDB'),
        top=Side(style='thin', color='D5DBDB'),
        bottom=Side(style='thin', color='D5DBDB')
    )
    
    # Alignements
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    # Fonctions utilitaires
    def merge_and_style(row, col_start, col_end, text, font, fill=None, alignment=center_align):
        if col_start == col_end:
            cell = ws.cell(row=row, column=col_start)
        else:
            ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
            cell = ws.cell(row=row, column=col_start)
        
        cell.value = text
        cell.font = font
        cell.alignment = alignment
        if fill:
            cell.fill = fill
        return cell

    def write_headers(headers, start_row):
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thick_border

    def write_row(data, start_row, is_alternate=False):
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=start_row, column=col_num, value=value)
            cell.font = data_font
            cell.border = thin_border
            
            # Alignement selon la colonne
            if col_num in [3, 5]:  # Grade et Type - centré
                cell.alignment = center_align
            else:  # Nom, Prénom, Établissement - à gauche
                cell.alignment = left_align
            
            # Couleur de fond alternée
            if is_alternate:
                cell.fill = alternate_fill

    # Récupérer les informations utilisateur
    profile = getattr(user, 'profile', None)
    etablissement = profile.etablissement.name if profile and profile.etablissement else "Non défini"
    role = user.groups.first().name if user.groups.exists() else "Aucun rôle"

    row = 1

    if mode == 'users':
            # Titre principal centré et fusionné
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
            title_cell = ws.cell(row=1, column=1, value="👨‍🏫 Tous les Profs / Assistants")
            title_cell.font = title_font
            title_cell.alignment = center_align
            
            row = 3  # Ligne vide après le titre
            
            # En-têtes exactement comme dans l'image mais adaptés aux utilisateurs
            headers = ['Nom', 'Prénom', 'Email', 'Établissement', 'Rôle', 'Sessions créées', 'Sessions validées']
            write_headers(headers, row)
            row += 1

            # Récupération et tri des utilisateurs
            users = User.objects.select_related('profile__etablissement').prefetch_related('groups').order_by('last_name', 'first_name')
            
            # Compteur pour les lignes alternées
            user_count = 0
            total_sessions_created = 0
            total_sessions_checked = 0

            for user in users:
                profile = getattr(user, 'profile', None)
                etab = profile.etablissement.name if profile and profile.etablissement else "Non défini"
                role = user.groups.first().name if user.groups.exists() else "Aucun rôle"
                sessions_created = Session.objects.filter(created_by=user).count()
                sessions_checked = Session.objects.filter(checked_by=user).count()
                
                # Cumul pour les statistiques
                total_sessions_created += sessions_created
                total_sessions_checked += sessions_checked
                
                # Écriture des données avec contenu centré
                data = [
                    user.last_name or "-",
                    user.first_name or "-",
                    user.email or "-",
                    etab,
                    role,
                    sessions_created,
                    sessions_checked,
                ]
                
                for col_num, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col_num, value=value)
                    cell.font = data_font
                    cell.border = thin_border
                    
                    # TOUT LE CONTENU CENTRÉ
                    cell.alignment = center_align
                    
                    # Couleur de fond alternée
                    if user_count % 2 == 1:
                        cell.fill = alternate_fill
                
                row += 1
                user_count += 1

            # Ligne de séparation avant le résumé
            row += 1
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="BDC3C7")
            row += 1
            
            # Ligne de résumé statistique
            summary_data = [
                "TOTAL",
                f"{user_count} membre(s)",
                "-",
                "-", 
                "-",
                f"{total_sessions_created}",
                f"{total_sessions_checked}"
            ]
            
            for col_num, value in enumerate(summary_data, 1):
                cell = ws.cell(row=row, column=col_num, value=value)
                cell.font = Font(bold=True, size=11, color=primary_color)
                cell.alignment = center_align
                cell.fill = accent_fill
                cell.border = thin_border

            # Statistiques supplémentaires
            row += 2
            stats_info = [
                f"📊 Statistiques générales :",
                f"• Moyenne sessions créées par membre : {total_sessions_created/user_count:.1f}" if user_count > 0 else "• Aucun membre",
                f"• Moyenne sessions validées par membre : {total_sessions_checked/user_count:.1f}" if user_count > 0 else "",
                f"• Total activité : {total_sessions_created + total_sessions_checked} actions"
            ]
            
            for info in stats_info:
                if info:  # Éviter les lignes vides
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
                    info_cell = ws.cell(row=row, column=1, value=info)
                    info_cell.font = Font(name="Calibri", size=10, color="7F8C8D")
                    info_cell.alignment = left_align
                    row += 1

            # Configuration des largeurs de colonnes selon l'image (plus larges pour lisibilité)
            ws.column_dimensions['A'].width = 25.3   # Nom
            ws.column_dimensions['B'].width = 21.7   # Prénom
            ws.column_dimensions['C'].width = 34.1   # Email
            ws.column_dimensions['D'].width = 19.1   # Établissement
            ws.column_dimensions['E'].width = 14.4   # Rôle
            ws.column_dimensions['F'].width = 14.1   # Sessions créées
            ws.column_dimensions['G'].width = 14.1   # Sessions validées

            # Hauteur des lignes identique à l'image
            for row_num in range(1, row + 1):
                ws.row_dimensions[row_num].height = 18

    elif mode == 'etablissement':
            # Styles pour reproduire l'apparence professionnelle
            etablissement_title_font = Font(name="Calibri", bold=True, size=16, color="2E5BBA")
            section_title_font = Font(name="Calibri", bold=True, size=12, color=black_text)
            
            # Remplissage pour les titres d'établissement
            etablissement_fill = PatternFill("solid", fgColor="E8F2FF")
            
            etablissements = Etablissement.objects.all().order_by('name')
            
            for etab_index, etab in enumerate(etablissements):
                if etab_index > 0:
                    row += 2  # Espace entre les établissements
                    
                # Titre de l'établissement avec style amélioré
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                etab_cell = ws.cell(row=row, column=1, value=f"🏫 {etab.name}")
                etab_cell.font = etablissement_title_font
                etab_cell.fill = etablissement_fill
                etab_cell.alignment = center_align
                etab_cell.border = thin_border
                row += 2

                # SECTION PROFS/ASSISTANTS EN PREMIER
                section_cell = ws.cell(row=row, column=1, value="👨‍🏫 Profs / Assistants")
                section_cell.font = section_title_font
                row += 1
                
                # En-têtes pour les profs/assistants
                headers_profs = ['Nom', 'Prénom', 'Email', 'Rôle', 'Sessions créées', 'Sessions validées']
                write_headers(headers_profs, row)
                row += 1
                
                # Données des profs/assistants triées alphabétiquement
                users = User.objects.filter(profile__etablissement=etab).order_by('last_name', 'first_name')
                user_count = 0
                for user in users:
                    role = user.groups.first().name if user.groups.exists() else "Aucun rôle"
                    sessions_created = Session.objects.filter(created_by=user).count()
                    sessions_checked = Session.objects.filter(checked_by=user).count()
                    write_row([
                        user.last_name or "-",
                        user.first_name or "-", 
                        user.email or "-",
                        role,
                        sessions_created,
                        sessions_checked
                    ], row, is_alternate=(user_count % 2 == 1))
                    row += 1
                    user_count += 1
                
                # Si aucun prof/assistant
                if user_count == 0:
                    no_data_cell = ws.cell(row=row, column=1, value="Aucun professeur/assistant enregistré")
                    no_data_cell.font = Font(name="Calibri", italic=True, size=10, color="7F8C8D")
                    no_data_cell.alignment = center_align
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                    row += 1

                row += 2  # Espace entre les sections

                # SECTION LICENCIÉS EN SECOND
                section_cell = ws.cell(row=row, column=1, value="🥋 Licenciés")
                section_cell.font = section_title_font
                row += 1
                
                # En-têtes pour les licenciés
                headers_licencies = ['Nom', 'Prénom', 'Grade', 'Date naissance', 'Participations', 'Dernière présence']
                write_headers(headers_licencies, row)
                row += 1
                
                # Données des licenciés triées alphabétiquement
                licencies = Licence.objects.filter(etablissement=etab).annotate(
                    nb_presences=Count('sessions'),
                    last_session_date=Max('sessions__date')
                ).order_by('nom', 'prenom')
                
                licencie_count = 0
                for licencie in licencies:
                    # Formatage de la date de naissance
                    date_naissance = licencie.date_naissance.strftime("%d/%m/%Y") if licencie.date_naissance else "-"
                    # Formatage de la dernière présence
                    derniere_presence = licencie.last_session_date.strftime("%d/%m/%Y") if licencie.last_session_date else "Aucune"
                    
                    write_row([
                        licencie.nom,
                        licencie.prenom,
                        licencie.grade,
                        date_naissance,
                        licencie.nb_presences,
                        derniere_presence
                    ], row, is_alternate=(licencie_count % 2 == 1))
                    row += 1
                    licencie_count += 1
                
                # Si aucun licencié
                if licencie_count == 0:
                    no_data_cell = ws.cell(row=row, column=1, value="Aucun licencié enregistré")
                    no_data_cell.font = Font(name="Calibri", italic=True, size=10, color="7F8C8D")
                    no_data_cell.alignment = center_align
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                    row += 1

                # Résumé pour l'établissement
                row += 1
                summary_cell = ws.cell(row=row, column=1, value=f"📊 Résumé : {user_count} prof(s)/assistant(s) • {licencie_count} licencié(s)")
                summary_cell.font = Font(name="Calibri", bold=True, size=11, color="2E5BBA")
                summary_cell.fill = PatternFill("solid", fgColor="F8FAFC")
                summary_cell.alignment = center_align
                summary_cell.border = thin_border
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                row += 1

            # Configuration des largeurs de colonnes optimisées
            ws.column_dimensions['A'].width = 18   # Nom
            ws.column_dimensions['B'].width = 18   # Prénom  
            ws.column_dimensions['C'].width = 25   # Email/Grade
            ws.column_dimensions['D'].width = 18   # Rôle/Date naissance
            ws.column_dimensions['E'].width = 15   # Sessions créées/Participations
            ws.column_dimensions['F'].width = 18   # Sessions validées/Dernière présence

    else:  # mode == 'all' - Version améliorée avec style professionnel
        # En-tête du document décalé d'une colonne (colonnes C à G)
        # Titre principal avec style amélioré
        merge_and_style(1, 3, 7, "EXPORT COMPLET", title_font, accent_fill)
        merge_and_style(2, 3, 7, "Licenciés et Utilisateurs", subtitle_font)

        # Informations contextuelles décalées (ligne 4)
        ws.cell(row=4, column=3, value="Exporté par :").font = info_font
        ws.cell(row=4, column=4, value=f"{user.get_full_name() or user.username}").font = data_font
        
        ws.cell(row=4, column=6, value="Date :").font = info_font
        ws.cell(row=4, column=6).alignment = right_align
        ws.cell(row=4, column=7, value=export_date).font = data_font
        ws.cell(row=4, column=7).alignment = right_align

        ws.cell(row=5, column=3, value="Rôle :").font = info_font
        ws.cell(row=5, column=4, value=role).font = data_font
        
        ws.cell(row=5, column=6, value="Heure :").font = info_font
        ws.cell(row=5, column=6).alignment = right_align
        ws.cell(row=5, column=7, value=export_time).font = data_font
        ws.cell(row=5, column=7).alignment = right_align

        ws.cell(row=6, column=3, value="Établissement :").font = info_font
        ws.cell(row=6, column=4, value=etablissement).font = data_font

        # En-têtes du tableau remontées (ligne 7) avec nouvelle colonne Participations
        header_row = 7
        headers = ['Nom', 'Prénom', 'Grade', 'Établissement', 'Participations', 'Type']
        
        # Écriture des en-têtes décalés d'une colonne (colonne B à G)
        for col_num, header in enumerate(headers, 2):  # Commence à la colonne B (2)
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thick_border
        
        row = 8  # Première ligne de données

        # Préparer les données pour tri uniforme
        all_persons = []
        
        # Récupérer tous les licenciés avec annotations
        licencies = Licence.objects.select_related('etablissement').annotate(
            nb_presences=Count('sessions')
        )
        
        for licencie in licencies:
            all_persons.append({
                'nom': licencie.nom,
                'prenom': licencie.prenom,
                'grade': licencie.grade,
                'etablissement': licencie.etablissement.name if licencie.etablissement else "Non défini",
                'participations': licencie.nb_presences,
                'type': "🥋 Licencié",
                'is_licencie': True
            })

        # Récupérer tous les utilisateurs
        users = User.objects.select_related('profile__etablissement')
        
        for user_obj in users:
            profile = getattr(user_obj, 'profile', None)
            etab = profile.etablissement.name if profile and profile.etablissement else "Non défini"
            user_role = user_obj.groups.first().name if user_obj.groups.exists() else "Aucun rôle"
            sessions_created = Session.objects.filter(created_by=user_obj).count()
            
            all_persons.append({
                'nom': user_obj.last_name or "-",
                'prenom': user_obj.first_name or "-",
                'grade': f"👨‍🏫 {user_role}",
                'etablissement': etab,
                'participations': sessions_created,  # Sessions créées pour les utilisateurs
                'type': "👤 Utilisateur",
                'is_licencie': False
            })

        # Tri alphabétique uniforme par nom puis prénom
        all_persons.sort(key=lambda x: (x['nom'].lower(), x['prenom'].lower()))

        # Compteurs
        licencie_count = sum(1 for p in all_persons if p['is_licencie'])
        user_count = sum(1 for p in all_persons if not p['is_licencie'])

        # Écriture des données avec lignes alternées
        for data_row_count, person in enumerate(all_persons):
            data = [
                person['nom'],
                person['prenom'],
                person['grade'],
                person['etablissement'],
                person['participations'],
                person['type']
            ]
            
            # Écriture décalée d'une colonne (colonne B à G)
            for col_offset, value in enumerate(data):
                col_num = col_offset + 2  # Décalage d'une colonne
                cell = ws.cell(row=row, column=col_num, value=value)
                cell.font = data_font
                cell.border = thin_border
                
                # Alignement selon la colonne
                if col_num in [4, 6, 7]:  # Grade, Participations, Type - centré
                    cell.alignment = center_align
                else:  # Nom, Prénom, Établissement - à gauche
                    cell.alignment = left_align
                
                # Couleur de fond alternée
                if data_row_count % 2 == 1:
                    cell.fill = alternate_fill
            
            row += 1

        # Ligne de séparation avant le résumé
        row += 1
        for col in range(2, 8):  # Décalé d'une colonne
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="BDC3C7")
        row += 1
        
        # Ligne de résumé statistique décalée
        summary_data = [
            "TOTAL",
            f"{len(all_persons)} personnes",
            "-",
            "-",
            f"{sum(p['participations'] for p in all_persons)}",
            f"{licencie_count} licenciés • {user_count} utilisateurs"
        ]
        
        for col_offset, value in enumerate(summary_data):
            col_num = col_offset + 2  # Décalage d'une colonne
            cell = ws.cell(row=row, column=col_num, value=value)
            cell.font = Font(bold=True, size=11, color=primary_color)
            cell.alignment = center_align if col_num in [2, 3, 6, 7] else left_align
            cell.fill = accent_fill
            cell.border = thin_border

        # Statistiques détaillées décalées
        row += 2
        merge_and_style(row, 3, 7, 
                       f"📊 Répartition : {licencie_count} licenciés et {user_count} utilisateurs", 
                       Font(name="Calibri", bold=True, size=11, color=secondary_color),
                       PatternFill("solid", fgColor="F8FAFC"))

        # Pied de page décalé
        footer_row = row + 3
        merge_and_style(footer_row, 3, 7, 
                       f"Document généré automatiquement le {export_date} à {export_time}", 
                       Font(size=8, italic=True, color=text_light))

        # Configuration des largeurs de colonnes optimisées avec décalage
        column_config = {
            'A': 2,    # Marge gauche
            'B': 18,   # Nom
            'C': 18,   # Prénom  
            'D': 22,   # Grade/Rôle
            'E': 18,   # Établissement
            'F': 15,   # Participations
            'G': 15,   # Type
            'H': 2     # Marge droite
        }

        for col_letter, width in column_config.items():
            ws.column_dimensions[col_letter].width = width

        # Propriétés de la page
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left = 0.7
        ws.page_margins.right = 0.7
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75

        # En-tête et pied de page pour l'impression
        ws.oddHeader.center.text = "Export Complet - Licenciés et Utilisateurs"
        ws.oddHeader.center.font = "Calibri,Bold"
        ws.oddFooter.center.text = "Page &P sur &N"

    # Finalisation
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

# Gestion prof
@login_required
@user_passes_test(lambda u: u.is_superuser)
def voir_utilisateurs(request):
    users = User.objects.prefetch_related('groups').filter(is_superuser=False).exclude(id=request.user.id)
    return render(request, 'presence/voir_utilisateurs.html', {'users': users})

@login_required
def user_detail(request, user_id):
    user_obj = get_object_or_404(User, pk=user_id)

    nb_sessions_creees = Session.objects.filter(created_by=user_obj).count()
    nb_sessions_checkees = Session.objects.filter(checked_by=user_obj).count()

    return render(request, 'presence/user_detail.html', {
        'user_obj': user_obj,
        'nb_sessions_creees': nb_sessions_creees,
        'nb_sessions_checkees': nb_sessions_checkees,
    })

@login_required
def ajouter_utilisateur_prof(request):
    if not request.user.is_superuser:
        messages.warning(request, "Accès refusé")
        return redirect('voir_utilisateurs')  # redirige vers la page d'accueil

    if request.method == 'POST':
        form = UserCreationWithGroupForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Prof ajouté avec succès.")
            return redirect('voir_utilisateurs')
    else:
        form = UserCreationWithGroupForm()

    return render(request, 'presence/ajouter_utilisateur.html', {'form': form})

@login_required
@user_passes_test(lambda u: u.is_superuser)
def user_edit(request, user_id):
    user_obj = get_object_or_404(User, pk=user_id)
    profile = getattr(user_obj, 'profile', None)

    if request.method == 'POST':
        form = UserUpdateForm(request.POST, instance=user_obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Utilisateur mis à jour avec succès.")
            return redirect('voir_utilisateurs')
    else:
        initial = {
            'etablissement': profile.etablissement if profile else None,
            'group': user_obj.groups.first()
        }
        form = UserUpdateForm(instance=user_obj, initial=initial)

    return render(request, 'presence/user_edit.html', {
        'form': form,
        'user_obj': user_obj
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def user_delete(request, user_id):
    user_obj = get_object_or_404(User, pk=user_id)

    if request.method == 'POST':
        # Optionnel : Empêcher un superuser de se supprimer lui-même
        if user_obj == request.user:
            messages.error(request, "Vous ne pouvez pas supprimer votre propre compte.")
            return redirect('voir_utilisateurs')

        user_obj.delete()
        messages.success(request, "Utilisateur supprimé avec succès.")
        return redirect('voir_utilisateurs')

    return render(request, 'presence/user_confirm_delete.html', {'user_obj': user_obj})

class SetPasswordForSelfView(LoginRequiredMixin, FormView):
    template_name = 'presence/set_password.html'
    form_class = SetPasswordForm
    success_url = reverse_lazy('user_detail')

    def get_form(self, form_class=None):
        return self.form_class(user=self.request.user, **self.get_form_kwargs())

    def form_valid(self, form):
        form.save()
        messages.success(self.request, "Mot de passe mis à jour avec succès.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['target_user'] = self.request.user  # Ajoute l'utilisateur au contexte
        return context

    def get_success_url(self):
        return reverse_lazy('user_detail', kwargs={'user_id': self.request.user.id})
    
class SetPasswordForOtherUserView(LoginRequiredMixin, UserPassesTestMixin, FormView):
    template_name = 'presence/set_password.html'
    form_class = SetPasswordForm

    def dispatch(self, request, *args, **kwargs):
        self.target_user = get_object_or_404(User, pk=self.kwargs['user_id'])
        return super().dispatch(request, *args, **kwargs)

    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.target_user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['target_user'] = self.target_user
        return context

    def form_valid(self, form):
        form.save()
        messages.success(self.request, f"Le mot de passe de {self.target_user.username} a été mis à jour.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('user_detail', kwargs={'user_id': self.target_user.id})

# gestion ville
@login_required
@user_passes_test(lambda u: u.is_superuser)
def gestion_ville(request):
    villes = Ville.objects.all()
    return render(request, 'presence/gestion_ville.html', {'villes': villes})

@login_required
@user_passes_test(lambda u: u.is_superuser)
def ajouter_ville(request):
    if request.method == 'POST':
        form = VilleForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Ville ajoutée avec succès.")
            return redirect('gestion_ville')
    else:
        form = VilleForm()
    return render(request, 'presence/ajouter_ville.html', {'form': form})

@user_passes_test(lambda u: u.is_superuser)
def modifier_ville(request, ville_id):
    ville = get_object_or_404(Ville, pk=ville_id)

    if request.method == 'POST':
        form = VilleForm(request.POST, instance=ville)
        if form.is_valid():
            form.save()
            messages.success(request, "Ville modifiée avec succès.")
            return redirect('gestion_ville')
    else:
        form = VilleForm(instance=ville)

    return render(request, 'presence/modifier_ville.html', {'form': form, 'ville': ville})


@login_required
@user_passes_test(lambda u: u.is_superuser)
def supprimer_ville(request, ville_id):
    ville = get_object_or_404(Ville, pk=ville_id)

    if request.method == 'POST':
        ville.delete()
        messages.success(request, "Ville supprimée avec succès.")
        return redirect('gestion_ville')

    return render(request, 'presence/supprimer_ville.html', {'ville': ville})


# gestion établissement
@login_required
@user_passes_test(lambda u: u.is_superuser)
def gestion_etablissement(request):
    etablissements = Etablissement.objects.select_related('ville').all()
    return render(request, 'presence/gestion_etablissement.html', {'etablissements': etablissements})

@login_required
@user_passes_test(lambda u: u.is_superuser)
def ajouter_etablissement(request):
    if request.method == 'POST':
        form = EtablissementForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Établissement ajouté avec succès.")
            return redirect('gestion_etablissement')
    else:
        form = EtablissementForm()
    return render(request, 'presence/ajouter_etablissement.html', {'form': form})

@login_required
@user_passes_test(lambda u: u.is_superuser)
def modifier_etablissement(request, etablissement_id):
    etablissement = get_object_or_404(Etablissement, pk=etablissement_id)

    if request.method == 'POST':
        form = EtablissementForm(request.POST, instance=etablissement)
        if form.is_valid():
            form.save()
            messages.success(request, "Établissement modifié avec succès.")
            return redirect('gestion_etablissement')
    else:
        form = EtablissementForm(instance=etablissement)

    return render(request, 'presence/modifier_etablissement.html', {'form': form, 'etablissement': etablissement})


@login_required
@user_passes_test(lambda u: u.is_superuser)
def supprimer_etablissement(request, etablissement_id):
    etablissement = get_object_or_404(Etablissement, pk=etablissement_id)

    if request.method == 'POST':
        etablissement.delete()
        messages.success(request, "Établissement supprimé avec succès.")
        return redirect('gestion_etablissement')

    return render(request, 'presence/supprimer_etablissement.html', {'etablissement': etablissement})

# gestion role
@login_required
@user_passes_test(lambda u: u.is_superuser)
def gestion_role(request):
    groupes = Group.objects.all()
    return render(request, 'presence/gestion_role.html', {'groupes': groupes})

@login_required
@user_passes_test(lambda u: u.is_superuser)
def ajouter_role(request):
    if request.method == 'POST':
        form = GroupForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Rôle (groupe) ajouté avec succès.")
            return redirect('gestion_role')
    else:
        form = GroupForm()

    return render(request, 'presence/ajouter_role.html', {'form': form})

@login_required
@user_passes_test(lambda u: u.is_superuser)
def modifier_role(request, role_id):
    group = get_object_or_404(Group, pk=role_id)

    if request.method == 'POST':
        form = GroupForm(request.POST, instance=group)
        if form.is_valid():
            form.save()
            messages.success(request, "Rôle mis à jour avec succès.")
            return redirect('gestion_role')
    else:
        form = GroupForm(instance=group)

    return render(request, 'presence/modifier_role.html', {'form': form, 'group': group})


@login_required
@user_passes_test(lambda u: u.is_superuser)
def supprimer_role(request, role_id):
    group = get_object_or_404(Group, pk=role_id)

    if request.method == 'POST':
        group.delete()
        messages.success(request, "Rôle supprimé avec succès.")
        return redirect('gestion_role')

    return render(request, 'presence/supprimer_role.html', {'group': group})